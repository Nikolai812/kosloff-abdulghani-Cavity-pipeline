import csv
import os
import configparser
from configparser import SectionProxy

from selenium import webdriver
from selenium.webdriver.common.by import By
from selenium.webdriver.support.ui import WebDriverWait, Select
from selenium.webdriver.support import expected_conditions as EC
from selenium.webdriver.chrome.service import Service
from selenium.webdriver.chrome.options import Options
import time

from file_namer import FileNamer, MethodType

def load_config():
    config = configparser.ConfigParser()
    config.read('config.ini')
    return config['DEFAULT']


def upload_and_submit_pdb(driver, pdb_file_path, pdb_input):
    try:
        # Wait for the dropdown to be present
        dropdown_locator = (By.ID, "cavityInputType")
        dropdown = WebDriverWait(driver, 10).until(
            EC.presence_of_element_located(dropdown_locator)
        )

        # Use Select to choose the "PDB File" option
        select = Select(dropdown)
        select.select_by_visible_text("PDB File")

        # Wait for the file input to appear
        file_input_locator = (By.CSS_SELECTOR, "input[type='file'][accept='pdb']")
        file_input = WebDriverWait(driver, 10).until(
            EC.presence_of_element_located(file_input_locator)
        )

        # Upload the PDB file
        file_input.send_keys(pdb_file_path)
        success_div_locator = (By.XPATH,
                               "//input[@type='file' and @accept='pdb']/following-sibling::div[text()='Success.']")
        submit_div = WebDriverWait(driver, 50).until(
            EC.element_to_be_clickable(success_div_locator)
        )

        print(f"Successfully uploaded the file: {pdb_input}")
        time.sleep(1)

        # Wait for the Submit button to be clickable
        submit_button_locator = (By.CSS_SELECTOR, "button.btn.btn-primary[type='submit']")
        submit_button = WebDriverWait(driver, 10).until(
            EC.element_to_be_clickable(submit_button_locator)
        )

        # Scroll the submit button into view
        driver.execute_script("arguments[0].scrollIntoView();", submit_button)
        WebDriverWait(driver, 10).until(EC.element_to_be_clickable(submit_button))
        # Use JavaScript to click the submit button
        driver.execute_script("arguments[0].click();", submit_button)
        print("Successfully clicked the Submit button")

    except Exception as e:
        print(f"An error occurred during upload and submit: {e}")
        raise


def prepare_cavity_tables(driver, pocket_limit=-1):
    """Extract cavity data and return tables as lists of rows."""
    try:
        # Wait for the Cavity Results table
        table_locator = (By.CSS_SELECTOR, "div.accordion-collapse.show table.table")
        table = WebDriverWait(driver, 30).until(
            EC.presence_of_element_located(table_locator)
        )

        # Locate ALL top-level rows
        rows = driver.find_elements(
            By.XPATH,
            "//button[.//b[normalize-space()='Download results']]"
            "/ancestor::div[contains(@class,'container')]"
            "//table[1]/tbody/tr[count(td)=7]"
        )
        total_rows = len(rows)
        print(f"Found {total_rows} cavity rows.")

        # Apply pocket limit if specified
        if pocket_limit > 0:
            if pocket_limit > total_rows:
                print(
                    f"WARNING: pocket_limit={pocket_limit} exceeds total cavities ({total_rows}). "
                    f"Processing all available cavities instead."
                )
                rows_to_process = rows
            else:
                print(f"Applying pocket limit: processing first {pocket_limit} cavities.")
                rows_to_process = rows[:pocket_limit]
        else:
            rows_to_process = rows

        # Initialize tables
        va_table = [['Cavity Number', 'Surface Area', 'Volume']]
        residues_table = [['Cavity Number', 'Chain', 'Seq ID', 'AA']]

        # Iterate over rows
        for cavity_index, row in enumerate(rows_to_process, start=1):
            print(f"\nProcessing cavity row {cavity_index}...")
            cavity_number = row.find_element(By.XPATH, "./td[1]").text.strip()
            if not cavity_number:
                raise RuntimeError(f"Could not read cavity number for row index {cavity_index}")

            # --- CLICK MORE BUTTON ---
            more_button = row.find_element(
                By.CSS_SELECTOR,
                "td:last-child div[style*='color: blue']"
            )
            driver.execute_script("arguments[0].scrollIntoView();", more_button)
            driver.execute_script("arguments[0].click();", more_button)
            print(f"Expanded details for row {cavity_index}")

            # Wait for expanded detail row
            detail_tr_id = f"more_{cavity_number}"
            WebDriverWait(driver, 15).until(
                EC.presence_of_element_located((By.ID, detail_tr_id))
            )
            detail_tr = driver.find_element(By.ID, detail_tr_id)
            WebDriverWait(driver, 10).until(
                lambda d: 'show' in detail_tr.get_attribute("class") or detail_tr.is_displayed()
            )

            # Extract data
            surface_area_td = detail_tr.find_element(
                By.XPATH,
                ".//th[contains(normalize-space(),'Surface Area')]/following-sibling::td"
            )
            volume_td = detail_tr.find_element(
                By.XPATH,
                ".//th[contains(normalize-space(),'Volume')]/following-sibling::td"
            )
            residues_td = detail_tr.find_element(
                By.XPATH,
                ".//th[contains(normalize-space(),'Residues')]/following-sibling::td"
            )

            surface_area = surface_area_td.text.strip()
            volume = volume_td.text.strip()
            residues_text = residues_td.text.strip()
            residues_list = [r.strip() for r in residues_text.split(',') if r.strip()]

            print(f"Surface Area: {surface_area}, Volume: {volume}, #residues: {len(residues_list)}")

            # --- POPULATE VA TABLE ---
            va_table.append([cavity_index, surface_area, volume])

            # --- POPULATE RESIDUES TABLE ---
            for residue in residues_list:
                aa, seq_id, chain = residue.split('-')
                residues_table.append([cavity_index, chain, seq_id, aa])

            # --- COLLAPSE ROW ---
            driver.execute_script("arguments[0].click();", more_button)
            print(f"Collapsed details for row {cavity_index}")

        print("\nAll cavities processed successfully.")
        return va_table, residues_table

    except Exception as e:
        print(f"An error occurred while preparing cavity tables: {e}")
        raise


def write_to_csv(va_table, residues_table, pdb_name, output_dir="output"):
    """Write tables to CSV files."""
    try:
        # Create output subfolder
        output_path = os.path.join(os.getcwd(), output_dir, pdb_name)
        os.makedirs(output_path, exist_ok=True)

        # File names
        va_csv_filename = FileNamer.get_va_name(pdb_name, MethodType.CVPL) + ".csv"
        residues_csv_filename = FileNamer.get_residues_name(pdb_name, MethodType.CVPL) + ".csv"

        # Write VA table
        with open(os.path.join(output_path, va_csv_filename), 'w', newline='', encoding='utf-8') as va_file:
            csv.writer(va_file).writerows(va_table)

        # Write residues table
        with open(os.path.join(output_path, residues_csv_filename), 'w', newline='', encoding='utf-8') as res_file:
            csv.writer(res_file).writerows(residues_table)

        print(f"CSV files written successfully to {output_path}")

    except Exception as e:
        print(f"An error occurred while writing CSV files: {e}")
        raise

def write_to_xlsx(va_table, residues_table, pdb_name, output_dir="output"):
    """Write tables to an Excel file with worksheets for VA data and per-cavity residues."""
    try:
        import openpyxl
        from openpyxl.utils import get_column_letter

        # Create output subfolder
        output_path = os.path.join(os.getcwd(), output_dir, pdb_name)
        os.makedirs(output_path, exist_ok=True)

        # File name
        xlsx_filename = FileNamer.get_residues_name(pdb_name, MethodType.CVPL) + ".xlsx"
        xlsx_path = os.path.join(output_path, xlsx_filename)

        # Create a new Excel workbook
        workbook = openpyxl.Workbook()

        # Remove the default sheet created by openpyxl
        if "Sheet" in workbook.sheetnames:
            workbook.remove(workbook["Sheet"])

        # --- WRITE VA TABLE TO 'Volumes and Areas' WORKSHEET ---
        va_sheet = workbook.create_sheet("Volumes and Areas")
        for row in va_table:
            va_sheet.append(row)

        # Auto-adjust column widths for VA sheet
        for column in va_sheet.columns:
            max_length = 0
            for cell in column:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            adjusted_width = (max_length + 2) * 1.2
            va_sheet.column_dimensions[get_column_letter(column[0].column)].width = adjusted_width

        # --- WRITE RESIDUES TABLE TO PER-CAVITY WORKSHEETS ---
        # Group residues by cavity number (first column in residues_table)
        cavities = {}
        for row in residues_table[1:]:  # Skip header
            cavity_num = row[0]
            if cavity_num not in cavities:
                cavities[cavity_num] = []
            cavities[cavity_num].append(row)  # Include cavity number in the data

        # Write each cavity's residues to a separate worksheet
        for cavity_num, residues in cavities.items():
            sheet_name = f"Cavity {cavity_num}"
            cavity_sheet = workbook.create_sheet(sheet_name)

            # Write header (including cavity number)
            cavity_sheet.append(residues_table[0])  # Full header from residues_table

            # Write residues
            for residue in residues:
                cavity_sheet.append(residue)

            # Auto-adjust column widths
            for column in cavity_sheet.columns:
                max_length = 0
                for cell in column:
                    try:
                        if len(str(cell.value)) > max_length:
                            max_length = len(str(cell.value))
                    except:
                        pass
                adjusted_width = (max_length + 2) * 1.2
                cavity_sheet.column_dimensions[get_column_letter(column[0].column)].width = adjusted_width

        # Save the workbook
        workbook.save(xlsx_path)
        print(f"Excel file written successfully to {xlsx_path}")

    except ImportError:
        raise ImportError("openpyxl is required for Excel export. Install it with: pip install openpyxl")
    except Exception as e:
        print(f"An error occurred while writing the Excel file: {e}")
        raise



def write_cavity_results(driver, pdb_name, output_dir="output", pocket_limit=-1):
    """Refactored method to prepare tables and write to CSV."""
    va_table, residues_table = prepare_cavity_tables(driver, pocket_limit)
    ## write_to_csv(va_table, residues_table, pdb_name, output_dir)
    write_to_xlsx(va_table, residues_table, pdb_name, output_dir)


def run_cavity_plus(pdb_input: str, config: SectionProxy):
    # Extract configuration values
    #config = load_config()
    chrome_driver_path = config['chrome_driver_path']
    cavity_plus_url = config['cavity_plus_url']
    input_dir = config['input_dir']
    output_dir = config['output_dir']
    pocket_limit = int(config['pocket_limit'])
    pdb_name=os.path.splitext(pdb_input)[0]

    # Construct the full path to the PDB file
    pdb_file_path = os.path.abspath(os.path.join(input_dir, pdb_input))

    # Set up Chrome options to automatically download files
    chrome_options = Options()
    script_dir = os.path.dirname(os.path.abspath(__file__))
    download_dir = os.path.join(script_dir, "output")
    prefs = {
        "download.default_directory": download_dir,
        "download.prompt_for_download": False,
        "download.directory_upgrade": True,
        "safebrowsing.enabled": False
    }
    chrome_options.add_experimental_option("prefs", prefs)

    # Set up the WebDriver with the specified path
    service = Service(chrome_driver_path)
    driver = webdriver.Chrome(service=service, options=chrome_options)

    try:
        # Open the CavityPlus URL
        driver.get(cavity_plus_url)

        # Upload and submit the PDB file
        upload_and_submit_pdb(driver, pdb_file_path, pdb_input)

        # Wait for the Download results button to appear and become clickable
        download_button_locator = (By.CSS_SELECTOR, "button.btn.btn-link b")
        download_button = WebDriverWait(driver, 300).until(
            EC.element_to_be_clickable(download_button_locator)
        )
        print("Download results button is now visible and clickable")

        write_cavity_results(driver, pdb_name, output_dir, pocket_limit=pocket_limit)

    except Exception as e:
        print(f"An error occurred: {e}")

    finally:
        # Close the browser
        print("This is finally, going to quit driver")
        driver.quit()

    print("Cavity Plus script completed")

if __name__ == '__main__':
    config = load_config()
    chrome_driver_path = config['chrome_driver_path']
    cavity_plus_url = config['cavity_plus_url']
    input_dir = config['input_dir']
    output_dir = config['output_dir']
    pdb_input = config['pdb_input']
    run_cavity_plus(pdb_input, config)
