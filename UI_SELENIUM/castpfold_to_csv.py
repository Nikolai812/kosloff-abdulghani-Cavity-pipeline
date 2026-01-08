import configparser
import csv
import logging
import openpyxl
import os

import time
from configparser import SectionProxy
from datetime import datetime

from selenium.webdriver.common.by import By
from selenium import webdriver
from selenium.webdriver.chrome.service import Service
from selenium.webdriver.chrome.options import Options
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC

from castpfold_request import submit_castpfold_request
from file_namer import FileNamer, MethodType

logger = logging.getLogger(__name__)

def enter_text_in_input(driver, input_id, text):
    """
    Locates an input field by ID and enters the specified text.

    Parameters:
    - input_id: The ID of the input field.
    - text: The text to enter into the input field.
    """
    input_field = driver.find_element(By.ID, input_id)
    input_field.clear()  # Clear any existing text
    input_field.send_keys(text)

def click_button_by_id(driver, button_id):
    """
    Locates a button by ID and clicks it.

    Parameters:
    - button_id: The ID of the button.
    """
    button = driver.find_element(By.ID, button_id)
    button.click()


def load_config():
    config = configparser.ConfigParser()
    config.read('config.ini')
    return config['DEFAULT']


def prepare_pocket_info_for_save(driver, pocket_limit):
    # Locate the table element
    table = driver.find_element(By.CSS_SELECTOR, "table")

    # Extract headers
    headers = []
    header_row = table.find_element(By.CSS_SELECTOR, "thead tr")
    for th in header_row.find_elements(By.TAG_NAME, "th"):
        if th.text.strip():  # Skip empty header cells
            headers.append(th.text.strip())

    # Extract rows
    rows = []
    tr_elements = table.find_elements(By.CSS_SELECTOR, "tbody tr")
    first_tr_elements = tr_elements[:pocket_limit]

    for tr in first_tr_elements:
        row = []
        for td in tr.find_elements(By.TAG_NAME, "td"):
            if td.text.strip():  # Skip empty cells
                row.append(td.text.strip())
        if row:  # Only add non-empty rows
            rows.append(row)

    return headers, rows

def prepare_atom_info_for_save(driver, pocket_limit):
    if pocket_limit > 10:
        raise ValueError(f"Pocket limit cannot be greater than 10, however requested pocket_limit was set to {pocket_limit}")

    # Locate all rows in the table
    pocket_rows = driver.find_elements(By.XPATH, "//table[.//th/div[contains(text(), 'Pocket ID')]]/tbody[1]/tr[contains(@class, 'ant-table-row-level-0')]")
    first_pocket_rows = pocket_rows[:pocket_limit]

    cav_list_all_atom_rows = []

    for i, pocket_row in enumerate(first_pocket_rows):
        try:
            # Get pocket ID from the second column
            pocket_id = pocket_row.find_elements(By.TAG_NAME, "td")[1].text.strip()

            # Click the expand icon
            expand_icon = pocket_row.find_element(By.CSS_SELECTOR, "td.ant-table-row-expand-icon-cell span.ant-table-row-collapsed")
            driver.execute_script("arguments[0].scrollIntoView({block: 'center', inline: 'center'});", expand_icon)
            WebDriverWait(driver, 10).until(EC.element_to_be_clickable(expand_icon))
            driver.execute_script("arguments[0].click();", expand_icon)
            time.sleep(1)  # Wait for the row to expand

            # Click the "Atom Info" element
            atom_info_header = pocket_row.find_element(By.XPATH, "./following-sibling::tr//*[contains(text(), 'Atom Info')]")
            atom_info_header.click()
            time.sleep(1)  # Wait for the atom info to load

            ul_atom_pagination = pocket_row.find_elements(By.XPATH, "./following-sibling::tr[contains(@class, 'ant-table-expanded-row-level-1') and not(contains(@style, 'display: none'))]//ul[contains(@class, 'ant-pagination')]")
            assert len(ul_atom_pagination) == 1, f"Expected exactly 1 pagination element, but found {len(ul_atom_pagination)} in {i} pocket row"

            li_pag_items = ul_atom_pagination[-1].find_elements(By.CSS_SELECTOR, "li.ant-pagination-item a")
            page_texts = [item.text for item in li_pag_items]
            atom_tab_count = int(page_texts[-1])

            # Initialize a set to store unique rows
            unique_atom_rows = set()

            # Click on the first pagination tab to begin
            first_button = ul_atom_pagination[-1].find_element(By.CSS_SELECTOR, "li.ant-pagination-item-1")
            driver.execute_script("arguments[0].scrollIntoView({block: 'center', inline: 'center'});", first_button)
            WebDriverWait(driver, 10).until(EC.element_to_be_clickable(first_button))
            driver.execute_script("arguments[0].click();", first_button)
            time.sleep(0.5)

            for ia in range(1, atom_tab_count + 1):
                # Extract rows from the current tab
                atom_info_table = pocket_row.find_element(By.XPATH, "./following-sibling::tr//div[contains(@class, 'ant-table-content')]//table")

                for tr in atom_info_table.find_elements(By.CSS_SELECTOR, "tbody tr"):
                    atom_row = []
                    for td in tr.find_elements(By.TAG_NAME, "td"):
                        if td.text.strip():
                            atom_row.append(td.text.strip())

                    if atom_row:
                        # Insert the cavity number (pocket_number = i + 1) as the first value in the row
                        atom_row.insert(0, str(i + 1))

                        # Remove the last (5th) value from the row
                        if len(atom_row) > 4:
                            atom_row = atom_row[:4]

                        # Add the row to the set as a tuple (to avoid duplicates)
                        unique_atom_rows.add(tuple(atom_row))

                logger.info(f"Atom pagination tab {ia} is displayed")

                if ia < atom_tab_count:
                    next_button = ul_atom_pagination[-1].find_element(By.CSS_SELECTOR, "li.ant-pagination-next a")
                    driver.execute_script("arguments[0].scrollIntoView({block: 'center', inline: 'center'});", next_button)
                    WebDriverWait(driver, 10).until(EC.element_to_be_clickable(next_button))
                    driver.execute_script("arguments[0].click();", next_button)
                    WebDriverWait(driver, 10).until(
                        EC.presence_of_element_located(
                            (By.CSS_SELECTOR, f"li.ant-pagination-item-{ia + 1}.ant-pagination-item-active"))
                    )
                    ul_atom_pagination = pocket_row.find_elements(By.XPATH, "./following-sibling::tr[contains(@class, 'ant-table-expanded-row-level-1') and not(contains(@style, 'display: none'))]//ul[contains(@class, 'ant-pagination')]")

            # Convert the set of tuples back to a list of lists
            all_atom_rows = [list(row) for row in unique_atom_rows]

            # Sort the rows by "Seq ID" (column index 2)
            all_atom_rows.sort(key=lambda x: int(x[2]))

            cav_list_all_atom_rows.append(all_atom_rows)

            # Close the "Atom Info" section
            time.sleep(1)
            atom_info_header = pocket_row.find_element(By.XPATH, "./following-sibling::tr//*[contains(text(), 'Atom Info')]")
            driver.execute_script("arguments[0].scrollIntoView({block: 'center', inline: 'center'});", atom_info_header)
            WebDriverWait(driver, 10).until(EC.element_to_be_clickable(atom_info_header))
            driver.execute_script("arguments[0].click();", atom_info_header)
            time.sleep(0.5)

            # Collapse the row
            expand_icon = pocket_row.find_element(By.CSS_SELECTOR, "td.ant-table-row-expand-icon-cell span.ant-table-row-expanded")
            driver.execute_script("arguments[0].scrollIntoView({block: 'center', inline: 'center'});", expand_icon)
            WebDriverWait(driver, 10).until(EC.element_to_be_clickable(expand_icon))
            driver.execute_script("arguments[0].click();", expand_icon)
            time.sleep(0.5)

        except Exception as e:

            logger.info(f"Error processing pocket {i}: {e} at {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}")
            raise

    return cav_list_all_atom_rows

def write_cav_all_atom_rows_to_csv(cav_list_all_atom_rows, output_directory, pdb_name):
    headers = ['Cavity Number', 'Chain', 'Seq ID', 'AA']

    for i, all_atom_rows in enumerate(cav_list_all_atom_rows):
        residues_csv_file_name = FileNamer.get_residues_name(pdb_name, MethodType.CSPF)
        with open(f"{output_directory}/{pdb_name}/{residues_csv_file_name}_{i + 1}.csv", 'w', newline='', encoding='utf-8') as csvfile:
            writer = csv.writer(csvfile)
            writer.writerow(headers)
            writer.writerows(all_atom_rows)

        print(f"Atom info for pocket {i + 1} saved to {residues_csv_file_name}_{i + 1}.csv")


def write_cav_all_atom_rows_to_excel(headers, rows, cav_list_all_atom_rows, output_directory, pdb_name):
    # Create a new Excel workbook
    workbook = openpyxl.Workbook()

    # Remove the default sheet created by openpyxl
    workbook.remove(workbook.active)

    # Create a worksheet for "Volumes Areas" as the first sheet
    volumes_areas_sheet = workbook.create_sheet(title="Volumes and Areas", index=0)

    # Write headers and rows to the "Volumes Areas" sheet
    volumes_areas_sheet.append(headers)
    for row in rows:
        volumes_areas_sheet.append(row)

    # Create a worksheet for each cavity
    for i, all_atom_rows in enumerate(cav_list_all_atom_rows):
        worksheet = workbook.create_sheet(title=f"Cavity {i + 1}")

        # Write headers
        worksheet.append(['Cavity Number', 'Chain', 'Seq ID', 'AA'])

        # Write data rows
        for row in all_atom_rows:
            worksheet.append(row)

    # Save the workbook to an Excel file
    residues_excel_file_name = FileNamer.get_residues_name(pdb_name, MethodType.CSPF)
    excel_file_path = f"{output_directory}/{pdb_name}/{residues_excel_file_name}.xlsx"
    workbook.save(excel_file_path)

    logger.info(f"All atom info saved to {excel_file_path} at {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}")

def write_pockets_to_csv(headers, rows, output_directory, pdb_name):
    # Create the output directory if it doesn't exist
    os.makedirs(f"{output_directory}/{pdb_name}", exist_ok=True)

    # Generate the CSV filename
    cspf_va_filename = FileNamer.get_va_name(pdb_name, MethodType.CSPF) + ".csv"

    # Write to CSV
    with open(f"{output_directory}/{pdb_name}/{cspf_va_filename}", 'w', newline='', encoding='utf-8') as csvfile:
        writer = csv.writer(csvfile)
        writer.writerow(headers)
        writer.writerows(rows)

    logger.info(f"Data written to {cspf_va_filename} at {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}")





def iterate_pagination(driver, output_directory, pdb_name: str, pocket_limit =1):
    if pocket_limit > 10:
        raise ValueError(f"Pocket limit cannot be greater than 10, however requested pocket_limit was set to {pocket_limit}")

    # Locate the pagination element
    pagination = driver.find_element(By.CSS_SELECTOR, "ul.ant-pagination")
    time.sleep(1)
    # Find all pagination items (excluding "Previous" and "Next" buttons)
    pagination_items = pagination.find_elements(By.CSS_SELECTOR, "li.ant-pagination-item a")

    # Get the last tab number from the list
    last_tab = pagination_items[-1]
    tab_count = int(last_tab.text)
    logger.info(f"Total pagination tabs for all pockets: {tab_count}")

    # In case of pocket limit < 10, only the first page tab should be treated
    max_pagination_item=2
    if(pocket_limit < 0):
        max_pagination_item=tab_count+1
    for i in range(1, max_pagination_item):
        # Wait for the current tab to be active
        WebDriverWait(driver, 10).until(
            EC.presence_of_element_located((By.CSS_SELECTOR, f"li.ant-pagination-item-{i}.ant-pagination-item-active"))
        )

        # Print the current tab number
        logger.info(f"Pagination tab {i} is displayed")

        # write_pocket_info_csv(driver, output_directory, pdb_name, pocket_limit)
        # open_atom_info_save_csv(driver, output_directory, pdb_name, pocket_limit)

        # Prepare data tables for saving
        cav_va_headers, cav_va_rows = prepare_pocket_info_for_save(driver, pocket_limit)
        cav_list_all_atom_rows = prepare_atom_info_for_save(driver, pocket_limit)
        #########

        # Save prepared data to file(s)
        # write_pockets_to_csv(cav_va_headers, cav_va_rows, output_directory, pdb_name)
        # write_cav_all_atom_rows_to_csv(cav_list_all_atom_rows, output_directory, pdb_name)
        write_cav_all_atom_rows_to_excel(cav_va_headers, cav_va_rows, cav_list_all_atom_rows, output_directory, pdb_name)
        ###############

        # If not the last tab, click "Next Page"
        if i < tab_count:
            next_button = pagination.find_element(By.CSS_SELECTOR, "li.ant-pagination-next a")
            next_button.click()
            # Wait for the next tab to load
            WebDriverWait(driver, 10).until(
                EC.presence_of_element_located((By.CSS_SELECTOR, f"li.ant-pagination-item-{i+1}.ant-pagination-item-active"))
            )

    # Return to the first tab
    first_tab = pagination.find_element(By.CSS_SELECTOR, "li.ant-pagination-item-1 a")
    first_tab.click()

    # Print completion message
    logger.info(f"All {max_pagination_item-1} tabs displayed since pocket limit is set: {pocket_limit} at {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}")


def run_castpfold(pdb_file, config: SectionProxy):
    logger.info("Starting CASTpFold script...  at {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}")
    # Load config
    # config = load_config()
    chrome_driver_path = config['chrome_driver_path']
    base_url = config['base_url']
    #job_number = config['job_number']
    output_directory = config['output_dir']
    #config['out_dir']  + '_' + job_number
    pocket_limit = int(config['pocket_limit'])
    # To be requested from the input directory using pdb_name
    logger.info(f"CASTpFold request for a file: {pdb_file}")
    input_dir = config['input_dir']
    if not FileNamer.verify_pdb_exists(input_dir, pdb_file):
        raise Exception(f"File {pdb_file} does not exist in the {input_dir}")
    job_number = submit_castpfold_request(os.path.join(input_dir, pdb_file))
    pdb_name = os.path.splitext(pdb_file)[0]
    logger.info(f"CASTpFold script initialized from config, pocket_limit: {pocket_limit}")

    # Set up Chrome options to automatically download files
    chrome_options = Options()
    script_dir = os.path.dirname(os.path.abspath(__file__))
    download_dir = os.path.join(script_dir, "output")
    prefs = {
        "download.default_directory": download_dir,
        "download.prompt_for_download": False,
        "download.directory_upgrade": True,
        "safebrowsing.enabled": False
    }
    chrome_options.add_experimental_option("prefs", prefs)

    # Set up the WebDriver with the specified path
    service = Service(chrome_driver_path)
    driver = webdriver.Chrome(service=service, options=chrome_options)

    try:
        # Open the specified URL
        logger.info(f"Waiting 20 secs for the job to complete...")
        time.sleep(20)
        logger.info(f"Loading castpFold page... at {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}")
        driver.get(f"{base_url}?{job_number}")
        # Wait until the button is visible and clickable
        download_button = WebDriverWait(driver, 20).until(
            EC.element_to_be_clickable(
                (By.XPATH, "//button[contains(@class, 'ant-btn-primary') and .//span[text()='Download CASTpFold Data']]")
            )
        )
        logger.info(f"Download button appeared with text: {download_button.text}")
        time.sleep(1)
        iterate_pagination(driver, output_directory=output_directory,pdb_name=pdb_name, pocket_limit=pocket_limit)
        time.sleep(1)
    except BaseException as e:
        logger.error(f"Error while running casfpfold  on {pdb_file} at {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}: ", e)
    finally:
        logger.info(f"castpfold finally, going to quit driver at {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}")
        driver.quit()

    logger.info("CASTpFold script completed.")


if __name__ == '__main__':
    config = load_config()
    #run_castpfold("input/HsOR343CF_1", config)